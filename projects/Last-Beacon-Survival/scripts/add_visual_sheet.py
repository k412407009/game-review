"""往 Last-Beacon review.xlsx 追加"视觉索引" Sheet (带缩略图)。

跑 generate_review.py 之后再跑这个。作用:
  1) 把 6 张 Google Play 独立商店截图 + 7 张关键视频帧都贴进 Excel
  2) 每张图有编号/类型/中文名/描述, 对应 review 文字里的"封面图/基地图/抽卡图..."和"scene_xxxx"
  3) 用户看 Issues sheet 里的"封面图" 三字, 翻到"视觉索引"一眼就能看到是哪张画面

依赖: openpyxl + Pillow (都已随 ppt-master 安装)

TODO (skill 层, 未来):
  * 把这个作为 generate_review.py 的 optional step (--with-visuals flag), 不用每次另跑
  * 让它支持读 <project>/raw_assets/<game>/store/*/*.jpg 的任意 store source (不只 googleplay)
  * 让 key_scenes 从 review.json 的 video_evidence.key_scenes_human_read 动态读取
"""
from __future__ import annotations

import io
import json
import shutil
import sys
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[attr-defined]
except Exception:
    pass

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage


PROJECT_DIR = Path(
    r"F:\Git\丁亮的个人助手\丁开心的游戏观察\external-game-reviews\Last-Beacon-Survival"
)
RAW_ROOT = PROJECT_DIR / "raw_assets" / "Last-Beacon-Survival"
STORE_DIR = RAW_ROOT / "store" / "googleplay"
LONG_VID_DIR = (
    RAW_ROOT
    / "gameplay"
    / "frames"
    / "last_beacon_survival_ACHEI_o_novo_rts_mobile_OxWu88bJ7Is"
)
SHORTS_DIR = (
    RAW_ROOT
    / "gameplay"
    / "frames"
    / "Last_Beacon_Survival_gaming_fyp_games_gameplay_2l4DO5Z10jo"
)
THUMB_DIR = PROJECT_DIR / "review" / ".thumbs"
XLSX_PATH = (
    PROJECT_DIR
    / "review"
    / "Last_Beacon_-_Survival_(外部游戏评审___Google_Play___com.hnhs.endlesssea.gp)_review.xlsx"
)

THICK_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

THUMB_MAX_W = 200
THUMB_MAX_H = 356  # 360x640 源图按比例缩到 200 宽 -> 356 高


# ========= 素材清单 =========

STORE_ITEMS: list[dict] = [
    {
        "code": "封面图",
        "src": STORE_DIR / "screenshot_01.jpg",
        "category": "Google Play 商店截图",
        "label": "封面图 (cover)",
        "desc": "灯塔孤岛情感主画面: 灯塔矗立海面孤岛 + 5 个幸存者围着篝火 + 远处乌云+鲨鱼环绕。Google Play 第 1/7/13 张是同一张画面 (重复 3 遍提高曝光)",
    },
    {
        "code": "基地图",
        "src": STORE_DIR / "screenshot_02.jpg",
        "category": "Google Play 商店截图",
        "label": "基地图",
        "desc": "主城俯视: 木栈道 + 灯塔 + 资源建筑群 (厨房/宿舍/锯木厂). 玩家日常最常看到的界面。Google Play 第 2/8/14 张",
    },
    {
        "code": "采集图",
        "src": STORE_DIR / "screenshot_03.jpg",
        "category": "Google Play 商店截图",
        "label": "采集图",
        "desc": "资源采集/探索: 小队到野外/岛屿采集. Google Play 第 3/9/15 张",
    },
    {
        "code": "防御战图",
        "src": STORE_DIR / "screenshot_04.jpg",
        "category": "Google Play 商店截图",
        "label": "防御战图",
        "desc": "基地防御: 敌方来袭, 玩家部署炮台/部队防守. Google Play 第 4/10/16 张",
    },
    {
        "code": "联盟战图",
        "src": STORE_DIR / "screenshot_05.jpg",
        "category": "Google Play 商店截图",
        "label": "联盟战图",
        "desc": "小队/联盟战: 战力 16.4K vs 19.1M 对比, 典型 4X SLG 联盟战画面. Google Play 第 5/11/17 张",
    },
    {
        "code": "抽卡图",
        "src": STORE_DIR / "screenshot_06.jpg",
        "category": "Google Play 商店截图",
        "label": "抽卡图",
        "desc": "英雄详情: Dylan S 品级, LV.23 战力 9,284, 1,125,841 钻石. 重度 SLG 抽卡+战力付费证据. Google Play 第 6/12/18 张",
    },
]

VIDEO_ITEMS: list[dict] = [
    {
        "code": "V1",
        "src": LONG_VID_DIR / "scene_1281.jpg",
        "category": "长视频关键帧 (葡语 8:05)",
        "label": "玩家档案 Lv.7 (0:43)",
        "desc": "[FLA]Survivor#1utp98h / 公会 [FLA]flamengo (巴西弗拉门戈俱乐部文化) / VIP Lv.1 — 确认巴西本地化深度绑定",
    },
    {
        "code": "V2",
        "src": LONG_VID_DIR / "scene_3396.jpg",
        "category": "长视频关键帧 (葡语 8:05)",
        "label": "主城岛屿全景 (1:53)",
        "desc": "Lv.7 主城已是岛屿全景 + 10+ 功能建筑 (Farol灯塔/Enfermaria医务室/Cozinha厨房/Serraria锯木厂/Dormitório宿舍/Acampamento de Aventura冒险营地) — 海洋主题递减的关键证据",
    },
    {
        "code": "V3",
        "src": LONG_VID_DIR / "scene_5289.jpg",
        "category": "长视频关键帧 (葡语 8:05)",
        "label": "抽卡双池 (2:56)",
        "desc": "Recrutamento Avançado 高级抽卡限免 4:32 + Recrutamento Lendário 传奇抽卡 23:59:38, 蓝发+红发女英雄立绘 — 双频次抽卡付费钩子实锤",
    },
    {
        "code": "V4",
        "src": LONG_VID_DIR / "scene_8196.jpg",
        "category": "长视频关键帧 (葡语 8:05)",
        "label": "基地近景 (4:33)",
        "desc": "木阁楼/渔网/帐篷/炮台/渔船/仓库, 英雄 16/16 满编, 11:46 保护罩倒计时 — 典型 SLG 基地布局, 海洋元素只体现在装饰层",
    },
    {
        "code": "V5",
        "src": LONG_VID_DIR / "scene_9771.jpg",
        "category": "长视频关键帧 (葡语 8:05)",
        "label": "遗物系统 Relíquias (5:25)",
        "desc": "Relíquias Perdidas 4 大类: Artesão 工匠 / Verdância 植物 / Zumbido 酒瓶 / Memórias 记忆. 叠加于抽卡之上的非卡牌养成层 — D5 付费深度的双轴证据",
    },
    {
        "code": "V6",
        "src": SHORTS_DIR / "scene_1527.jpg",
        "category": "Shorts 关键帧 (印尼语 3:00)",
        "label": "新手木筏基地 (6s)",
        "desc": "Dapur Lv.1 厨房 50% 建造中, 0/1 工人槽位 — 早期木筏期 (主题最强时段) 画面, 跟 Lv.7 全景对比落差巨大",
    },
    {
        "code": "V7",
        "src": SHORTS_DIR / "scene_2252.jpg",
        "category": "Shorts 关键帧 (印尼语 3:00)",
        "label": "Bab 1 任务列表 (12s)",
        "desc": "Bab 1 'Bertahan di Laut 在海上生存' 章节, 4 条任务 (升篝火 Lv.1 → 重建厨房 → 重建渔夫小屋 → 升篝火 Lv.2), 主角持火把立绘 — 新手期海洋主题最强证据",
    },
]


def make_thumb(src: Path, dst: Path) -> None:
    """生成 PNG 缩略图 (PNG 避免 JPG 重压缩, 同时 openpyxl 吃 PNG 最稳)。"""
    if dst.exists() and dst.stat().st_mtime >= src.stat().st_mtime:
        return
    img = PILImage.open(src)
    img.thumbnail((THUMB_MAX_W, THUMB_MAX_H), PILImage.LANCZOS)
    if img.mode != "RGB":
        img = img.convert("RGB")
    dst.parent.mkdir(parents=True, exist_ok=True)
    img.save(dst, format="PNG", optimize=True)


def main() -> int:
    if not XLSX_PATH.exists():
        print(f"ERROR: xlsx not found: {XLSX_PATH}")
        return 2

    THUMB_DIR.mkdir(parents=True, exist_ok=True)

    items = STORE_ITEMS + VIDEO_ITEMS

    # 1) 生成缩略图
    thumb_paths: list[Path] = []
    for it in items:
        if not it["src"].exists():
            print(f"  WARN: missing source {it['src']}")
            thumb_paths.append(Path(""))
            continue
        name = f"{it['code']}__{it['src'].stem}.png"
        tp = THUMB_DIR / name
        make_thumb(it["src"], tp)
        thumb_paths.append(tp)
    print(f"  thumbs ready: {THUMB_DIR} ({len(thumb_paths)} files)")

    # 2) 打开 xlsx, 加 sheet "视觉索引"
    wb = load_workbook(XLSX_PATH)
    if "视觉索引" in wb.sheetnames:
        del wb["视觉索引"]
    ws = wb.create_sheet("视觉索引")

    # ---- 表头 ----
    headers = ["编号", "类型", "画面", "描述", "缩略图"]
    col_widths = [10, 28, 22, 60, 32]
    for c, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF", name="Microsoft YaHei", size=11)
        cell.fill = PatternFill("solid", fgColor="1A2C5C")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THICK_BORDER
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 28

    # ---- 介绍行 (合并单元格说明) ----
    intro = (
        "对 Excel 其他 Sheet 里出现的 '封面图/基地图/采集图/防御战图/联盟战图/抽卡图' 以及 "
        "'V1-V7' 的视觉对照。所有图来自 Google Play 商店 (共 18 张, 6 张独立 × 3 遍重复) + "
        "YouTube 两个 gameplay 视频 (葡语 8:05 + 印尼语 Shorts 3:00, 总 80 帧 AI 打标后选 7 张关键帧)。"
    )
    ws.cell(row=2, column=1, value=intro)
    ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=5)
    cell = ws.cell(row=2, column=1)
    cell.font = Font(italic=True, color="555555", name="Microsoft YaHei", size=10)
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    cell.fill = PatternFill("solid", fgColor="FFF7E6")
    ws.row_dimensions[2].height = 48

    # ---- 数据行 + 嵌图 ----
    row = 3
    for it, tp in zip(items, thumb_paths):
        ws.cell(row=row, column=1, value=it["code"])
        ws.cell(row=row, column=2, value=it["category"])
        ws.cell(row=row, column=3, value=it["label"])
        ws.cell(row=row, column=4, value=it["desc"])
        ws.cell(row=row, column=5, value="" if tp.exists() else "(图源缺失)")

        for c in range(1, 6):
            cell = ws.cell(row=row, column=c)
            cell.font = Font(name="Microsoft YaHei", size=10)
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left" if c == 4 else "center")
            cell.border = THICK_BORDER

        # 加粗 "画面" 列
        ws.cell(row=row, column=3).font = Font(name="Microsoft YaHei", size=11, bold=True, color="1A2C5C")

        # 高亮商店 vs 视频的分组色
        if it in STORE_ITEMS:
            for c in range(1, 6):
                ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor="E8F2FF")
        else:
            for c in range(1, 6):
                ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor="FFF5E8")

        # 嵌图 (如果存在)
        if tp.exists():
            img = XLImage(str(tp))
            # openpyxl 默认 96 DPI 下, 行高单位 pt. 1 pt ≈ 1.333 px. 图高 ~260 px -> 行高 195
            img.anchor = f"E{row}"
            ws.add_image(img)
            ws.row_dimensions[row].height = 196
        else:
            ws.row_dimensions[row].height = 24

        row += 1

    # ---- Sheet 排到"Scores"后面 (第 3 位) ----
    desired_order = ["Issues", "Scores", "视觉索引", "Action_Items"]
    wb._sheets = [wb[n] for n in desired_order if n in wb.sheetnames]  # type: ignore[attr-defined]

    wb.save(XLSX_PATH)
    print(f"  wrote: {XLSX_PATH}")
    print(f"  sheet '视觉索引' rows: {row - 1} items (6 store + 7 video)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
