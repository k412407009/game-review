"""
Review · 视觉索引 Sheet 生成器 (通用版)

往 <project>/review/<project>_review.xlsx 追加 "视觉索引" Sheet, 带缩略图。
为什么需要这个 sheet: 评审 Excel 里会出现 "封面图 / 基地图 / V1 / V2" 这种短标签,
没有视觉索引的话非技术读者根本不知道对应哪张画面。

两种入口:
  1. 作为模块被 generate_review.py --with-visuals 调用 (推荐)
  2. 作为独立 CLI: python add_visual_sheet.py <project_dir>

数据源 (按优先级):
  * 商店截图:
     - 优先读 review.json 的 visual_catalog.store (完整 label/desc)
     - 没有就自动扫描 <project>/raw_assets/**/store/**/*.{jpg,jpeg,png},
       生成默认条目 (label = "商店图 N", desc 为空)
  * 视频关键帧:
     - 从 review.json 的 video_evidence.key_scenes_human_read 读
     - desc 缺失时, 回退到 raw_assets/**/gameplay/descriptions.json
     - frame 字段如 "scene_1281 (长视频 43s)", 自动在
       <project>/raw_assets/**/gameplay/frames/**/scene_1281.* 里反查实际文件路径
     - 即使图源缺失, 也保留该视频行并明确标 "(图源缺失)"
     - 如果整个 review.json 没有 video_evidence, 则跳过视频部分 (例如内部 PPT 评审)

依赖: openpyxl, Pillow (PIL)

usage:
  python add_visual_sheet.py <project_dir>              # 默认自动找 xlsx
  python add_visual_sheet.py <project_dir> --xlsx X.xlsx
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[attr-defined]
except Exception:
    pass

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage


THICK_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# 缩略图尺寸: 200 宽足够在 Excel 里看清, 高度按原图比例自适应 (最高 356)
THUMB_MAX_W = 200
THUMB_MAX_H = 356

# 颜色
COLOR_STORE_BG = "E8F2FF"   # 淡蓝 (商店图)
COLOR_VIDEO_BG = "FFF5E8"   # 淡橙 (视频帧)
COLOR_HEADER_BG = "1A2C5C"  # 深蓝 (表头)
COLOR_INTRO_BG = "FFF7E6"   # 浅黄 (说明行)

STORE_SOURCES = ("googleplay", "appstore", "steam", "tapTap", "huawei", "others")

SCENE_ID_RE = re.compile(r"scene_(\d+)", re.IGNORECASE)


# ============== 辅助 ==============

def _msg(quiet: bool, text: str) -> None:
    if not quiet:
        print(text)


def _make_thumb(src: Path, dst: Path) -> None:
    """按原图比例压到 ≤200x356 的 PNG 缩略图 (PNG 防二次压缩, openpyxl 兼容最好)。"""
    if dst.exists() and dst.stat().st_mtime >= src.stat().st_mtime:
        return
    img = PILImage.open(src)
    img.thumbnail((THUMB_MAX_W, THUMB_MAX_H), PILImage.LANCZOS)
    if img.mode != "RGB":
        img = img.convert("RGB")
    dst.parent.mkdir(parents=True, exist_ok=True)
    img.save(dst, format="PNG", optimize=True)


def _find_xlsx(project_dir: Path) -> Path | None:
    """在 <project_dir>/review/ 里找 *_review.xlsx (取最新)。"""
    review_dir = project_dir / "review"
    if not review_dir.exists():
        return None
    candidates = sorted(review_dir.glob("*_review.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0] if candidates else None


def _find_review_json(project_dir: Path) -> Path | None:
    review_dir = project_dir / "review"
    if not review_dir.exists():
        return None
    candidates = sorted(review_dir.glob("*_review.json"))
    return candidates[0] if candidates else None


# ============== 数据源: 商店截图 ==============

def _discover_store_images(project_dir: Path, review_data: dict[str, Any]) -> list[dict]:
    """
    发现商店截图.
    优先级 1: review_data['visual_catalog']['store'] (完整 label/desc)
    优先级 2: 自动扫 <project>/raw_assets/*/store/*/*.{jpg,jpeg,png}, 给默认 label

    返回 [{code, src(Path), category, label, desc}]
    """
    catalog = review_data.get("visual_catalog", {}).get("store")
    if catalog:
        items: list[dict] = []
        for entry in catalog:
            raw_path = entry.get("path") or entry.get("src")
            if not raw_path:
                continue
            src = (project_dir / raw_path).resolve()
            items.append({
                "code": entry.get("code", entry.get("label", src.stem)),
                "src": src,
                "category": entry.get("category", "商店截图"),
                "label": entry.get("label", src.stem),
                "desc": entry.get("desc", ""),
            })
        return items

    items: list[dict] = []
    raw_root = project_dir / "raw_assets"
    if not raw_root.exists():
        return items

    for game_dir in sorted(raw_root.iterdir()):
        if not game_dir.is_dir():
            continue
        store_root = game_dir / "store"
        if not store_root.exists():
            continue
        for source_dir in sorted(store_root.iterdir()):
            if not source_dir.is_dir():
                continue
            files = sorted([
                p for p in source_dir.glob("*")
                if p.suffix.lower() in (".jpg", ".jpeg", ".png", ".webp")
            ])
            seen_stems: set[str] = set()
            idx = 0
            for f in files:
                if f.stem in seen_stems:
                    continue
                seen_stems.add(f.stem)
                idx += 1
                items.append({
                    "code": f"商店图 {idx}",
                    "src": f,
                    "category": f"{source_dir.name} 商店截图",
                    "label": f"{source_dir.name} 商店第 {idx} 张",
                    "desc": f"(来自 raw_assets/{game_dir.name}/store/{source_dir.name}/{f.name}. review.json 未提供描述, 用默认)",
                })
    return items


# ============== 数据源: 视频关键帧 ==============

def _load_descriptions(project_dir: Path) -> tuple[dict[str, str], dict[str, str]]:
    """聚合 raw_assets/**/gameplay/descriptions.json, 返回相对路径/scene_id 两种索引。"""
    raw_root = project_dir / "raw_assets"
    by_rel: dict[str, str] = {}
    by_scene: dict[str, str] = {}
    if not raw_root.exists():
        return by_rel, by_scene

    for game_dir in sorted(raw_root.iterdir()):
        if not game_dir.is_dir():
            continue
        desc_path = game_dir / "gameplay" / "descriptions.json"
        if not desc_path.exists():
            continue
        try:
            payload = json.loads(desc_path.read_text(encoding="utf-8"))
        except Exception:
            continue
        for rel, desc in payload.items():
            desc = str(desc).strip()
            if not desc or desc.startswith("ERROR"):
                continue
            by_rel.setdefault(rel, desc)
            m = SCENE_ID_RE.search(rel)
            if m:
                by_scene.setdefault(m.group(1), desc)
    return by_rel, by_scene


def _guess_missing_frame_path(project_dir: Path, scene_id: str) -> Path | None:
    raw_root = project_dir / "raw_assets"
    if not raw_root.exists():
        return None
    game_dirs = [d for d in sorted(raw_root.iterdir()) if d.is_dir()]
    if not game_dirs:
        return None
    return game_dirs[0] / "gameplay" / "frames" / "__missing__" / f"scene_{scene_id}.jpg"


def _find_frame_file(project_dir: Path, scene_id: str) -> Path | None:
    """
    在 <project>/raw_assets/**/gameplay/frames/**/scene_<id>.* 里找帧文件。
    scene_id 可能带前缀 0 也可能不带 (scene_1281 / scene_01281)。
    """
    raw_root = project_dir / "raw_assets"
    if not raw_root.exists():
        return None

    variants = {scene_id, scene_id.lstrip("0") or "0", scene_id.zfill(4), scene_id.zfill(5)}
    for game_dir in sorted(raw_root.iterdir()):
        if not game_dir.is_dir():
            continue
        frames_root = game_dir / "gameplay" / "frames"
        if not frames_root.exists():
            continue
        for sub in frames_root.iterdir():
            if not sub.is_dir():
                continue
            for ext in ("jpg", "jpeg", "png"):
                for v in variants:
                    cand = sub / f"scene_{v}.{ext}"
                    if cand.exists():
                        return cand
    return None


def _discover_video_scenes(project_dir: Path, review_data: dict[str, Any]) -> list[dict]:
    """从 review_data.video_evidence.key_scenes_human_read 组装视频关键帧条目。"""
    ve = review_data.get("video_evidence") or {}
    frame_analysis = ve.get("frame_analysis") or {}
    scenes = frame_analysis.get("key_scenes_human_read") or []
    _, desc_by_scene = _load_descriptions(project_dir)

    items: list[dict] = []
    for idx, s in enumerate(scenes, 1):
        frame_str: str = s.get("frame", "")
        m = SCENE_ID_RE.search(frame_str)
        if not m:
            continue
        scene_num = m.group(1)
        path = _find_frame_file(project_dir, scene_num)
        if path is None:
            path = _guess_missing_frame_path(project_dir, scene_num)

        dims = s.get("dims_affected") or []
        dims_tag = f" [影响 {', '.join(dims)}]" if dims else ""
        content = (s.get("content") or "").strip()
        if not content:
            content = desc_by_scene.get(scene_num, "")

        items.append({
            "code": f"V{idx}",
            "src": path,
            "category": frame_str,
            "label": f"关键帧 V{idx}",
            "desc": f"{content}{dims_tag}",
        })
    return items


# ============== 主逻辑: 建 sheet ==============

def _build_header(ws, store_count: int, video_count: int) -> None:
    headers = ["编号", "类型", "画面", "描述", "缩略图"]
    col_widths = [10, 28, 22, 60, 32]
    for c, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF", name="Microsoft YaHei", size=11)
        cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THICK_BORDER
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 28

    store_clause = f"{store_count} 张商店截图" if store_count else "无商店截图"
    video_clause = f"{video_count} 张视频关键帧" if video_count else "无视频证据"
    intro = (
        f"本 Sheet 是对 Excel 其他 Sheet 里出现的视觉标签 (封面图/基地图/V1/V2 等) 的"
        f"图文对照。\n自动索引范围: {store_clause} + {video_clause}。\n"
        f"商店截图来自项目 raw_assets/*/store/*, 视频关键帧来自项目 video_evidence.key_scenes_human_read。"
    )
    ws.cell(row=2, column=1, value=intro)
    ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=5)
    cell = ws.cell(row=2, column=1)
    cell.font = Font(italic=True, color="555555", name="Microsoft YaHei", size=10)
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    cell.fill = PatternFill("solid", fgColor=COLOR_INTRO_BG)
    ws.row_dimensions[2].height = 62


def _build_row(ws, row: int, item: dict, thumb_path: Path | None, is_store: bool) -> None:
    ws.cell(row=row, column=1, value=item["code"])
    ws.cell(row=row, column=2, value=item["category"])
    ws.cell(row=row, column=3, value=item["label"])
    ws.cell(row=row, column=4, value=item["desc"])
    ws.cell(row=row, column=5, value="" if thumb_path and thumb_path.exists() else "(图源缺失)")

    bg = COLOR_STORE_BG if is_store else COLOR_VIDEO_BG
    for c in range(1, 6):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name="Microsoft YaHei", size=10)
        cell.alignment = Alignment(
            wrap_text=True,
            vertical="top",
            horizontal="left" if c == 4 else "center",
        )
        cell.border = THICK_BORDER
        cell.fill = PatternFill("solid", fgColor=bg)

    # 加粗画面列, 与 code 区分
    ws.cell(row=row, column=3).font = Font(
        name="Microsoft YaHei", size=11, bold=True, color="1A2C5C"
    )

    if thumb_path and thumb_path.exists():
        img = XLImage(str(thumb_path))
        img.anchor = f"E{row}"
        ws.add_image(img)
        ws.row_dimensions[row].height = 196
    else:
        ws.row_dimensions[row].height = 24


def add_visual_sheet(
    project_dir: Path,
    *,
    xlsx_path: Path | None = None,
    quiet: bool = False,
) -> int:
    """
    主入口. 给 <project>/review/*_review.xlsx 追加 "视觉索引" Sheet.
    返回: 新 sheet 的数据行数 (0 表示没素材, 未新建 sheet)
    """
    if xlsx_path is None:
        xlsx_path = _find_xlsx(project_dir)
    if xlsx_path is None or not xlsx_path.exists():
        _msg(quiet, f"  WARN: xlsx not found under {project_dir}/review/, skip visual sheet.")
        return 0

    json_path = _find_review_json(project_dir)
    review_data: dict[str, Any] = {}
    if json_path and json_path.exists():
        try:
            review_data = json.loads(json_path.read_text(encoding="utf-8"))
        except Exception as e:
            _msg(quiet, f"  WARN: cannot parse review.json ({e}), fallback to blank.")

    store_items = _discover_store_images(project_dir, review_data)
    video_items = _discover_video_scenes(project_dir, review_data)

    if not store_items and not video_items:
        _msg(quiet, "  visual catalog empty (no store images & no video scenes) — skip visual sheet.")
        return 0

    thumb_dir = project_dir / "review" / ".thumbs"
    thumb_dir.mkdir(parents=True, exist_ok=True)

    def _prep_thumbs(items: list[dict]) -> list[Path | None]:
        out: list[Path | None] = []
        for it in items:
            src: Path | None = it.get("src")
            if src is None or not src.exists():
                if src is not None:
                    _msg(quiet, f"  WARN: missing {src}")
                out.append(None)
                continue
            name = f"{it['code']}__{src.stem}.png".replace(" ", "_")
            dst = thumb_dir / name
            try:
                _make_thumb(src, dst)
                out.append(dst)
            except Exception as e:
                _msg(quiet, f"  WARN: thumb fail {src} ({e})")
                out.append(None)
        return out

    store_thumbs = _prep_thumbs(store_items)
    video_thumbs = _prep_thumbs(video_items)

    _msg(
        quiet,
        f"  visual catalog: {len(store_items)} store + {len(video_items)} video (thumbs → {thumb_dir.name}/)",
    )

    # 打开 xlsx
    wb = load_workbook(xlsx_path)
    if "视觉索引" in wb.sheetnames:
        del wb["视觉索引"]
    ws = wb.create_sheet("视觉索引")

    _build_header(ws, len(store_items), len(video_items))

    row = 3
    for it, tp in zip(store_items, store_thumbs):
        _build_row(ws, row, it, tp, is_store=True)
        row += 1
    for it, tp in zip(video_items, video_thumbs):
        _build_row(ws, row, it, tp, is_store=False)
        row += 1

    desired_order = ["Issues", "Scores", "视觉索引", "Action_Items"]
    existing = wb.sheetnames
    wb._sheets = [wb[n] for n in desired_order if n in existing] + [  # type: ignore[attr-defined]
        wb[n] for n in existing if n not in desired_order
    ]

    wb.save(xlsx_path)
    data_rows = row - 3
    _msg(quiet, f"  wrote: {xlsx_path} (视觉索引: {data_rows} rows)")
    return data_rows


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="生成/追加 Excel 视觉索引 Sheet")
    parser.add_argument("project_dir", type=Path, help="项目目录 (含 review/ 子目录)")
    parser.add_argument("--xlsx", type=Path, default=None, help="显式指定 xlsx (默认自动查找)")
    parser.add_argument("--quiet", action="store_true", help="静默模式, 不打 info 日志")
    args = parser.parse_args(argv)

    project_dir = args.project_dir.resolve()
    if not project_dir.exists():
        print(f"ERROR: project_dir not found: {project_dir}")
        return 2

    add_visual_sheet(project_dir, xlsx_path=args.xlsx, quiet=args.quiet)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
